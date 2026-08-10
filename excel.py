from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule, CellIsRule


# ── Palette ──────────────────────────────────────────────────────────────────
C_NAVY    = "1B2A4A"   # headers principal
C_BLUE    = "2E4FA3"   # sub-headers
C_TEAL    = "1A7B6B"   # accent / positive
C_RED     = "C0392B"   # negative / alert
C_GOLD    = "D4A017"   # auspicios
C_WHITE   = "FFFFFF"
C_LGRAY   = "F4F6FA"
C_MGRAY   = "DDE2EC"
C_INPUT   = "EBF5FB"   # editable cells
C_FORMULA = "000000"   # formula cells → black text

# ── Helpers ──────────────────────────────────────────────────────────────────
def hdr(text, wb=None):
    return Font(name="Arial", bold=True, color=C_WHITE, size=11)

def cell_font(bold=False, color=C_FORMULA, size=10):
    return Font(name="Arial", bold=bold, color=color, size=size)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def center(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def left(wrap=False):
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)

def right_align():
    return Alignment(horizontal="right", vertical="center")

THIN = Side(style="thin", color="BBBBBB")
MED  = Side(style="medium", color="888888")

def thin_border():
    return Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def med_border():
    return Border(left=MED, right=MED, top=MED, bottom=MED)

CLP = '#,##0" CLP"'
CLP_NEG = '#,##0" CLP";(#,##0" CLP");"-"'

def style_header_row(ws, row, cols, bg=C_NAVY, text=C_WHITE, sz=11):
    for c in cols:
        cell = ws.cell(row=row, column=c)
        cell.font = Font(name="Arial", bold=True, color=text, size=sz)
        cell.fill = fill(bg)
        cell.alignment = center(wrap=True)
        cell.border = thin_border()

def style_subheader(ws, row, cols, bg=C_BLUE):
    for c in cols:
        cell = ws.cell(row=row, column=c)
        cell.font = Font(name="Arial", bold=True, color=C_WHITE, size=10)
        cell.fill = fill(bg)
        cell.alignment = center(wrap=True)
        cell.border = thin_border()

def style_data_row(ws, row, cols, bg=C_WHITE, alt=False):
    bg = C_LGRAY if alt else C_WHITE
    for c in cols:
        cell = ws.cell(row=row, column=c)
        cell.fill = fill(bg)
        cell.border = thin_border()

def style_total_row(ws, row, cols, bg=C_MGRAY):
    for c in cols:
        cell = ws.cell(row=row, column=c)
        cell.font = Font(name="Arial", bold=True, color=C_FORMULA, size=10)
        cell.fill = fill(bg)
        cell.alignment = right_align()
        cell.border = thin_border()

def set_input(ws, row, col, value=None, fmt=None):
    cell = ws.cell(row=row, column=col)
    if value is not None:
        cell.value = value
    cell.font = Font(name="Arial", color="0000CD", size=10)
    cell.fill = fill(C_INPUT)
    cell.border = thin_border()
    cell.alignment = right_align()
    if fmt:
        cell.number_format = fmt

def set_formula(ws, row, col, formula, fmt=None, bold=False, color=C_FORMULA):
    cell = ws.cell(row=row, column=col)
    cell.value = formula
    cell.font = Font(name="Arial", bold=bold, color=color, size=10)
    cell.border = thin_border()
    cell.alignment = right_align()
    if fmt:
        cell.number_format = fmt

# ════════════════════════════════════════════════════════════════════════════
wb = Workbook()

# ════════════════════════════════════════════════════════════════════════════
#  HOJA 1: PRESUPUESTO DE GASTOS
# ════════════════════════════════════════════════════════════════════════════
ws_g = wb.active
ws_g.title = "💰 Gastos"
ws_g.sheet_view.showGridLines = False
ws_g.freeze_panes = "A4"

# Column widths
widths = {"A": 5, "B": 32, "C": 28, "D": 12, "E": 14, "F": 16, "G": 22}
for col, w in widths.items():
    ws_g.column_dimensions[col].width = w
ws_g.row_dimensions[1].height = 38
ws_g.row_dimensions[2].height = 22
ws_g.row_dimensions[3].height = 30

# Title banner
ws_g.merge_cells("A1:G1")
t = ws_g["A1"]
t.value = "🎓 CONGRESO ELECTRÓNICO 2026 — PRESUPUESTO DE GASTOS"
t.font = Font(name="Arial", bold=True, color=C_WHITE, size=14)
t.fill = fill(C_NAVY)
t.alignment = center()

ws_g.merge_cells("A2:G2")
t2 = ws_g["A2"]
t2.value = "IEEE RAS UdeC  ·  Centro de Estudiantes Ing. Civil Electrónica  ·  Universidad de Concepción  ·  11–13 noviembre 2026"
t2.font = Font(name="Arial", color=C_WHITE, size=9)
t2.fill = fill(C_BLUE)
t2.alignment = center()

# Column headers row 3
ws_g["A3"] = "#"
ws_g["B3"] = "Categoría / Ítem de Gasto"
ws_g["C3"] = "Descripción / Detalle"
ws_g["D3"] = "Cantidad"
ws_g["E3"] = "Costo Unitario (CLP)"
ws_g["F3"] = "Total Ítem (CLP)"
ws_g["G3"] = "Notas"
style_header_row(ws_g, 3, range(1, 8))

# ── Expense data ─────────────────────────────────────────────────────────────
# Structure: (categoria, item, descripcion, cantidad, costo_unitario, nota)
gastos = [
    # --- Catering y Alimentación ---
    ("CATERING Y ALIMENTACIÓN", None, None, None, None, None),
    (None, "Coffee Break (días 1 y 2)", "Café, jugos, galletas, frutas — 2 coffee breaks/día × 2 días = 4 breaks", 120, 2500, "Estimado por persona, 120 asistentes promedio"),
    (None, "Coffee Break (día 3 — cierre)", "Break especial día de clausura", 120, 3000, "Puede incluir algo más elaborado para el cierre"),
    (None, "Almuerzo asistentes (días 2 y 3)", "Colación o voucher de almuerzo — 2 jornadas completas", 120, 4500, "Cotizar casino UdeC u opción externa"),
    (None, "Catering ponentes / organizadores", "Almuerzo diferenciado para expositores, talleristas y comité", 20, 8000, "~15 ponentes + 5 organizadores"),
    (None, "Celebración de cierre", "Bebidas, snacks y bocados para la actividad recreativa final", 100, 3500, "Estimado 100 personas en el cierre"),
    # --- Impresión y Materiales ---
    ("IMPRESIÓN Y MATERIALES", None, None, None, None, None),
    (None, "Pendones auspiciadores — Platino (XL)", "Pendón 80×200 cm con logo destacado, uno por auspiciador Platino", 3, 35000, "Cotizar imprenta local — pedido mínimo 3"),
    (None, "Pendones auspiciadores — Oro", "Pendón 60×160 cm, uno por auspiciador Oro", 2, 28000, ""),
    (None, "Pendones auspiciadores — Plata", "Pendón 60×160 cm, uno por auspiciador Plata", 1, 28000, ""),
    (None, "Pendones institucionales (IEEE/UdeC)", "Pendones generales del evento y co-organizadores", 4, 30000, "IEEE RAS, CEICE, UdeC, Congreso"),
    (None, "Credenciales / fotochecks", "Impresión + porta-credencial. Roles: asistente, expositor, org., voluntario", 200, 800, ""),
    (None, "Señalética del evento", "Indicadores de salas, stands numerados, zonas de coffee break", 30, 2500, "Incluir flechas y carteles de ubicación"),
    (None, "Carpetas / kits para asistentes", "Folder con material informativo, programa y bloc de notas", 150, 2200, ""),
    (None, "Flyers y material de difusión (físico)", "Afiches para campus UdeC y otras universidades", 100, 600, "Impresión A3 color"),
    (None, "Diplomas — ponentes y talleristas", "Diploma impreso, firmado y enmarcado (opcional)", 35, 1800, "~20 ponentes + 15 talleristas"),
    (None, "Diplomas — voluntarios", "Diploma oficial acreditación curricular para voluntarios", 25, 1200, ""),
    # --- Merchandising ---
    ("MERCHANDISING", None, None, None, None, None),
    (None, "Poleras equipo organizador", "Polera oficial del congreso para comité y voluntarios", 30, 8500, "Cotizar desde 20 unidades para precio especial"),
    (None, "Stickers / pins del evento", "Stickers con logo del congreso para asistentes", 300, 300, ""),
    (None, "Libretas / lápices (kit asistente)", "Material de trabajo para asistentes — incluible en carpeta", 150, 1500, ""),
    # --- Producción Audiovisual ---
    ("PRODUCCIÓN AUDIOVISUAL", None, None, None, None, None),
    (None, "Arriendo sistema de sonido (si falta)", "Micrófono inalámbrico extra, parlantes. Verificar equipo UdeC primero", 1, 80000, "Solo si el auditorio UdeC no lo provee"),
    (None, "Arriendo proyector extra (salas paralelas)", "Proyector para salas de talleres", 2, 40000, "Salas paralelas pueden no tener proyector"),
    (None, "Fotografía profesional (días 2 y 3)", "Fotógrafo para cobertura de las jornadas principales", 2, 60000, "Por jornada. Considerar estudiante de periodismo"),
    (None, "Transmisión / streaming (opcional)", "Cámara + software para transmitir algunas charlas en vivo", 1, 50000, "Opcional según presupuesto disponible"),
    (None, "Iluminación extra (escenario/auditorio)", "Luces de escenario para charlas plenarias", 1, 30000, "Cotizar con proveedores AV"),
    # --- Plataforma de Tickets ---
    ("PLATAFORMA Y COMISIONES", None, None, None, None, None),
    (None, "Comisión plataforma de tickets (PassLine/otro)", "Comisión promedio ~5-8% sobre ventas brutas de entradas", 1, 80000, "Ajustar según proyección de ventas en hoja Entradas"),
    (None, "Punto de venta físico (lector de tarjeta)", "Arriendo o comisión por terminal de pago para venta en puerta", 1, 15000, ""),
    # --- Materiales Talleres ---
    ("MATERIALES PARA TALLERES", None, None, None, None, None),
    (None, "Materiales taller de soldadura", "Soldadores, estaño, placas PCB, componentes", 20, 3500, "Por participante — taller de ~20 personas"),
    (None, "Materiales taller PCB", "Software (licencias o acceso), placas prototipo, insumos", 20, 2000, ""),
    (None, "Materiales taller instrumento electrónico", "Componentes específicos del instrumento por participante", 20, 4000, "Definir según tallerista confirmado"),
    (None, "Materiales talleres software (C, Matlab, LaTeX)", "Acceso a PC o salas de computación — costo de uso", 1, 20000, "Gestionar sala computación UdeC (puede ser $0)"),
    # --- Marketing Digital ---
    ("MARKETING Y DISEÑO DIGITAL", None, None, None, None, None),
    (None, "Diseño gráfico (externo, si aplica)", "Apoyo profesional para identidad visual si no lo hace el equipo", 1, 120000, "Puede ser $0 si lo hace equipo interno"),
    (None, "Pauta en redes sociales (ads Instagram/LinkedIn)", "Publicidad pagada para alcanzar más público", 1, 50000, "Recomendado para masificar convocatoria"),
    (None, "Dominio web + hosting (si aplica)", "Sitio web del congreso (opcional)", 1, 25000, "Alternativa gratuita: Linktree o perfil IG"),
    # --- Logística General ---
    ("LOGÍSTICA GENERAL", None, None, None, None, None),
    (None, "Transporte ponentes externos (desde otras ciudades)", "Pasajes + viáticos para expositores de otras regiones", 3, 60000, "Estimado 3 ponentes externos"),
    (None, "Alojamiento ponentes externos (si aplica)", "1 noche hotel para ponente de fuera de Concepción", 2, 65000, "Negociar con ponentes — muchos pueden no requerirlo"),
    (None, "Gastos administrativos (copias, estampillas, etc.)", "Trámites, documentos, convenios, sobres", 1, 20000, ""),
    (None, "Imprevistos y contingencia (10%)", "Reserva para gastos no contemplados", 1, 0, "⚠️ Este valor se calcula automáticamente abajo"),
]

row = 4
cat_rows = []   # (row, label) for category subtotal rows
expense_rows = []  # rows with actual expense data

for entry in gastos:
    cat, item, desc, qty, unit, nota = entry
    if cat:
        # Category header
        ws_g.row_dimensions[row].height = 22
        ws_g.merge_cells(f"A{row}:G{row}")
        c = ws_g[f"A{row}"]
        c.value = f"  {cat}"
        c.font = Font(name="Arial", bold=True, color=C_WHITE, size=10)
        c.fill = fill(C_BLUE)
        c.alignment = left()
        c.border = thin_border()
        cat_rows.append(row)
        row += 1
    else:
        # Data row
        ws_g.row_dimensions[row].height = 20
        alt = (row % 2 == 0)
        bg = C_LGRAY if alt else C_WHITE

        # A: index (auto)
        idx_cell = ws_g.cell(row=row, column=1)
        idx_cell.value = len(expense_rows) + 1
        idx_cell.font = Font(name="Arial", color="888888", size=9)
        idx_cell.fill = fill(bg)
        idx_cell.border = thin_border()
        idx_cell.alignment = center()

        # B: item name
        b = ws_g.cell(row=row, column=2)
        b.value = item
        b.font = Font(name="Arial", color=C_FORMULA, size=10)
        b.fill = fill(bg)
        b.border = thin_border()
        b.alignment = left(wrap=True)

        # C: descripcion
        c_cell = ws_g.cell(row=row, column=3)
        c_cell.value = desc
        c_cell.font = Font(name="Arial", color="555555", size=9)
        c_cell.fill = fill(bg)
        c_cell.border = thin_border()
        c_cell.alignment = left(wrap=True)

        # D: cantidad (input, blue)
        if item and "contingencia" in item.lower():
            # This row is auto-computed
            set_formula(ws_g, row, 4, 1, fmt='#,##0')
        else:
            set_input(ws_g, row, 4, qty, fmt='#,##0')

        # E: costo unitario (input, blue)
        if item and "contingencia" in item.lower():
            # Will be set later after we know subtotals
            set_formula(ws_g, row, 5, 0, fmt=CLP)
        else:
            set_input(ws_g, row, 5, unit, fmt=CLP)

        # F: total (formula, black)
        f_cell = ws_g.cell(row=row, column=6)
        f_cell.value = f"=D{row}*E{row}"
        f_cell.font = Font(name="Arial", bold=True, color=C_FORMULA, size=10)
        f_cell.fill = fill(bg)
        f_cell.border = thin_border()
        f_cell.alignment = right_align()
        f_cell.number_format = CLP

        # G: notes
        g_cell = ws_g.cell(row=row, column=7)
        g_cell.value = nota
        g_cell.font = Font(name="Arial", color="777777", size=9, italic=True)
        g_cell.fill = fill(bg)
        g_cell.border = thin_border()
        g_cell.alignment = left(wrap=True)

        expense_rows.append(row)
        row += 1

# ── Find contingencia row and fix its formula ─────────────────────────────
# Contingencia is last expense row
cont_row = expense_rows[-1]
non_cont_rows = expense_rows[:-1]
# Sum of all non-contingencia F cells * 10%
sum_refs = "+".join([f"F{r}" for r in non_cont_rows])
ws_g.cell(row=cont_row, column=5).value = f"=({sum_refs})*0.10"
ws_g.cell(row=cont_row, column=5).number_format = CLP
ws_g.cell(row=cont_row, column=5).font = Font(name="Arial", bold=True, color=C_FORMULA, size=10)
ws_g.cell(row=cont_row, column=5).fill = fill(C_LGRAY)

# ── TOTAL GASTOS row ──────────────────────────────────────────────────────
total_row = row
ws_g.row_dimensions[total_row].height = 26
ws_g.merge_cells(f"A{total_row}:E{total_row}")
tr = ws_g[f"A{total_row}"]
tr.value = "TOTAL PRESUPUESTO DE GASTOS"
tr.font = Font(name="Arial", bold=True, color=C_WHITE, size=11)
tr.fill = fill(C_NAVY)
tr.alignment = right_align()
tr.border = med_border()

total_f = ws_g.cell(row=total_row, column=6)
total_f.value = f"=SUM(F4:F{total_row-1})"
total_f.font = Font(name="Arial", bold=True, color=C_WHITE, size=12)
total_f.fill = fill(C_NAVY)
total_f.border = med_border()
total_f.alignment = right_align()
total_f.number_format = CLP

ws_g.cell(row=total_row, column=7).fill = fill(C_NAVY)
ws_g.cell(row=total_row, column=7).border = med_border()

TOTAL_GASTOS_REF = f"F{total_row}"
GASTOS_SHEET_TOTAL_ROW = total_row

# ── Legend box ───────────────────────────────────────────────────────────
leg_row = total_row + 2
ws_g.merge_cells(f"B{leg_row}:D{leg_row}")
ws_g[f"B{leg_row}"].value = "🔵 Celda azul = valor editable (ingresa aquí tus datos)"
ws_g[f"B{leg_row}"].font = Font(name="Arial", color="0000CD", size=9, italic=True)
ws_g[f"B{leg_row}"].alignment = left()
ws_g.merge_cells(f"E{leg_row}:G{leg_row}")
ws_g[f"E{leg_row}"].value = "⚫ Celda negra = fórmula automática (no editar)"
ws_g[f"E{leg_row}"].font = Font(name="Arial", color=C_FORMULA, size=9, italic=True)
ws_g[f"E{leg_row}"].alignment = left()

leg_row2 = leg_row + 1
ws_g.merge_cells(f"B{leg_row2}:G{leg_row2}")
ws_g[f"B{leg_row2}"].value = "⚠️  La Contingencia (10%) se calcula automáticamente sobre el subtotal sin contingencia. Los gastos en 0 son opcionales."
ws_g[f"B{leg_row2}"].font = Font(name="Arial", color=C_RED, size=9, italic=True)
ws_g[f"B{leg_row2}"].alignment = left()


# ════════════════════════════════════════════════════════════════════════════
#  HOJA 2: AUSPICIOS
# ════════════════════════════════════════════════════════════════════════════
ws_a = wb.create_sheet("🤝 Auspicios")
ws_a.sheet_view.showGridLines = False
ws_a.freeze_panes = "A5"

# Column widths
for col, w in {"A": 5, "B": 28, "C": 18, "D": 16, "E": 16,
               "F": 22, "G": 18, "H": 20}.items():
    ws_a.column_dimensions[col].width = w

ws_a.row_dimensions[1].height = 38
ws_a.row_dimensions[2].height = 22
ws_a.row_dimensions[3].height = 38
ws_a.row_dimensions[4].height = 28

# Title
ws_a.merge_cells("A1:H1")
t = ws_a["A1"]
t.value = "🤝 REGISTRO DE AUSPICIADORES — CONGRESO ELECTRÓNICO 2026"
t.font = Font(name="Arial", bold=True, color=C_WHITE, size=14)
t.fill = fill(C_NAVY)
t.alignment = center()

ws_a.merge_cells("A2:H2")
t2 = ws_a["A2"]
t2.value = "Agrega cada auspiciador y su monto — el Resumen se actualiza automáticamente"
t2.font = Font(name="Arial", color=C_WHITE, size=9)
t2.fill = fill(C_BLUE)
t2.alignment = center()

# Category reference table
ws_a.merge_cells("A3:H3")
ref = ws_a["A3"]
ref.value = ("Categorías de auspicio:    🥇 Platino = $1.000.000    🥈 Oro = $800.000    "
             "🥉 Plata = $600.000    🔵 Bronce = $400.000    🎁 Apoyo = Especie/Otro")
ref.font = Font(name="Arial", bold=True, color="3D2B00", size=10)
ref.fill = PatternFill("solid", fgColor="FFF3CD")
ref.alignment = center()
ref.border = thin_border()

# Column headers row 4
headers_a = ["#", "Empresa / Organización", "Categoría", "Monto Acordado (CLP)",
             "Monto Recibido (CLP)", "Beneficios comprometidos", "Estado", "Observaciones"]
for i, h in enumerate(headers_a, 1):
    c = ws_a.cell(row=4, column=i)
    c.value = h
    c.font = Font(name="Arial", bold=True, color=C_WHITE, size=10)
    c.fill = fill(C_GOLD if i in (3, 4, 5) else C_NAVY)
    c.alignment = center(wrap=True)
    c.border = thin_border()

# ── Sponsor data rows (15 editable rows) ─────────────────────────────────
SPONSOR_START = 5
SPONSOR_END = 24   # 20 rows for sponsors

# Category color map
CAT_FILLS = {
    "Platino": "FFF8E1",
    "Oro":     "FFFDE7",
    "Plata":   "F3F3F3",
    "Bronce":  "FFF3E0",
    "Apoyo":   "E8F5E9",
}

ejemplo_sponsors = [
    ("Empresa ABC Electrónica", "Platino", 1000000, 1000000, "Pendón XL, 3 pases VIP, charla técnica", "Recibido", "Pago completo recibido"),
    ("TechCorp Chile", "Oro", 800000, 0, "Pendón, 2 pases VIP, mención en redes", "Pendiente pago", "Firmó convenio el 15 may"),
]

for i in range(SPONSOR_START, SPONSOR_END + 1):
    row_i = i
    ws_a.row_dimensions[row_i].height = 20
    alt = (i % 2 == 0)
    bg = C_LGRAY if alt else C_WHITE

    # A: index
    idx = ws_a.cell(row=row_i, column=1)
    idx.value = i - SPONSOR_START + 1
    idx.font = Font(name="Arial", color="888888", size=9)
    idx.fill = fill(bg)
    idx.border = thin_border()
    idx.alignment = center()

    # B-H: editable cells
    ex_idx = i - SPONSOR_START
    for col in range(2, 9):
        c = ws_a.cell(row=row_i, column=col)
        c.fill = fill(C_INPUT)
        c.border = thin_border()
        c.font = Font(name="Arial", color="0000CD", size=10)
        c.alignment = left(wrap=True)
        if col in (4, 5):
            c.number_format = CLP
            c.alignment = right_align()

    # Prefill examples
    if ex_idx < len(ejemplo_sponsors):
        e = ejemplo_sponsors[ex_idx]
        ws_a.cell(row=row_i, column=2).value = e[0]
        ws_a.cell(row=row_i, column=3).value = e[1]
        ws_a.cell(row=row_i, column=4).value = e[2]
        ws_a.cell(row=row_i, column=5).value = e[3]
        ws_a.cell(row=row_i, column=6).value = e[4]
        ws_a.cell(row=row_i, column=7).value = e[5]
        ws_a.cell(row=row_i, column=8).value = e[6]

# ── Totales auspicios ─────────────────────────────────────────────────────
tot_a_row = SPONSOR_END + 1
ws_a.row_dimensions[tot_a_row].height = 26

ws_a.merge_cells(f"A{tot_a_row}:C{tot_a_row}")
ta = ws_a[f"A{tot_a_row}"]
ta.value = "TOTAL AUSPICIOS ACORDADOS / RECIBIDOS"
ta.font = Font(name="Arial", bold=True, color=C_WHITE, size=11)
ta.fill = fill(C_NAVY)
ta.alignment = right_align()
ta.border = med_border()

# Total acordado
t_ac = ws_a.cell(row=tot_a_row, column=4)
t_ac.value = f"=SUM(D{SPONSOR_START}:D{SPONSOR_END})"
t_ac.font = Font(name="Arial", bold=True, color=C_WHITE, size=11)
t_ac.fill = fill(C_NAVY)
t_ac.border = med_border()
t_ac.alignment = right_align()
t_ac.number_format = CLP

# Total recibido
t_rec = ws_a.cell(row=tot_a_row, column=5)
t_rec.value = f"=SUM(E{SPONSOR_START}:E{SPONSOR_END})"
t_rec.font = Font(name="Arial", bold=True, color=C_WHITE, size=11)
t_rec.fill = fill(C_NAVY)
t_rec.border = med_border()
t_rec.alignment = right_align()
t_rec.number_format = CLP

for col in range(6, 9):
    ws_a.cell(row=tot_a_row, column=col).fill = fill(C_NAVY)
    ws_a.cell(row=tot_a_row, column=col).border = med_border()

TOTAL_AUSPICIOS_ACORDADO_REF = f"'🤝 Auspicios'!D{tot_a_row}"
TOTAL_AUSPICIOS_RECIBIDO_REF = f"'🤝 Auspicios'!E{tot_a_row}"

# ── Legend ────────────────────────────────────────────────────────────────
leg = tot_a_row + 2
ws_a.merge_cells(f"A{leg}:H{leg}")
ws_a[f"A{leg}"].value = ("💡 Instrucciones: Ingresa el nombre de la empresa, selecciona la categoría, "
                          "el monto acordado en convenio y el monto efectivamente recibido. "
                          "Deja en 0 el monto recibido si aún no ha llegado el pago.")
ws_a[f"A{leg}"].font = Font(name="Arial", color="333333", size=9, italic=True)
ws_a[f"A{leg}"].alignment = left(wrap=True)
ws_a.row_dimensions[leg].height = 30


# ════════════════════════════════════════════════════════════════════════════
#  HOJA 3: PROYECCIÓN DE ENTRADAS
# ════════════════════════════════════════════════════════════════════════════
ws_e = wb.create_sheet("🎟️ Entradas")
ws_e.sheet_view.showGridLines = False
ws_e.freeze_panes = "A4"

for col, w in {"A": 5, "B": 30, "C": 14, "D": 14, "E": 14, "F": 14, "G": 16, "H": 18}.items():
    ws_e.column_dimensions[col].width = w

ws_e.row_dimensions[1].height = 38
ws_e.row_dimensions[2].height = 22

ws_e.merge_cells("A1:H1")
t = ws_e["A1"]
t.value = "🎟️ PROYECCIÓN DE INGRESOS POR ENTRADAS — CONGRESO ELECTRÓNICO 2026"
t.font = Font(name="Arial", bold=True, color=C_WHITE, size=13)
t.fill = fill(C_NAVY)
t.alignment = center()

ws_e.merge_cells("A2:H2")
t2 = ws_e["A2"]
t2.value = "Modifica los precios y cantidades estimadas — el total se actualiza automáticamente"
t2.font = Font(name="Arial", color=C_WHITE, size=9)
t2.fill = fill(C_BLUE)
t2.alignment = center()

# Column headers
hdrs_e = ["#", "Categoría de Asistente", "Precio P1\n(CLP)", "Precio P2\n(CLP)",
          "Precio P3\n(CLP)", "Precio Puerta\n(CLP)", "Asistentes\nesperados", "Ingreso\nProyectado (CLP)"]
for i, h in enumerate(hdrs_e, 1):
    c = ws_e.cell(row=3, column=i)
    c.value = h
    c.font = Font(name="Arial", bold=True, color=C_WHITE, size=10)
    c.fill = fill(C_TEAL)
    c.alignment = center(wrap=True)
    c.border = thin_border()
ws_e.row_dimensions[3].height = 35

# Ticket data: (categoría, p1, p2, p3, puerta, estimado)
tickets = [
    ("Estudiante IEEE activo", 7000, 9000, 12000, 15000, 30),
    ("Estudiante mechón UdeC (1er año)", 8000, 10000, 13000, 16000, 20),
    ("Estudiante UdeC (sin membresía IEEE)", 9000, 11000, 14000, 17000, 40),
    ("Estudiante otra universidad", 10000, 12000, 15000, 18000, 20),
    ("Docente / académico", 12000, 14000, 17000, 20000, 10),
    ("Profesional del área", 15000, 18000, 20000, 25000, 10),
    ("Expositor de paper (precio fijo)", 12000, 12000, 12000, 12000, 10),
    ("Invitado / cortesía (auspiciadores)", 0, 0, 0, 0, 5),
]

# Weights for price mix (P1 30%, P2 40%, P3 20%, Puerta 10%)
# Ingreso proyectado = asistentes * (p1*0.30 + p2*0.40 + p3*0.20 + puerta*0.10)

ent_rows = []
for i, (cat, p1, p2, p3, pue, est) in enumerate(tickets):
    r = 4 + i
    ws_e.row_dimensions[r].height = 20
    alt = (r % 2 == 0)
    bg = C_LGRAY if alt else C_WHITE

    # Index
    c0 = ws_e.cell(row=r, column=1)
    c0.value = i + 1
    c0.font = Font(name="Arial", color="888888", size=9)
    c0.fill = fill(bg); c0.border = thin_border(); c0.alignment = center()

    # Category name
    c1 = ws_e.cell(row=r, column=2)
    c1.value = cat
    c1.font = Font(name="Arial", color=C_FORMULA, size=10)
    c1.fill = fill(bg); c1.border = thin_border(); c1.alignment = left()

    # Prices (input)
    for ci, price in enumerate([p1, p2, p3, pue], 3):
        set_input(ws_e, r, ci, price, fmt=CLP)

    # Asistentes esperados (input)
    set_input(ws_e, r, 7, est, fmt='#,##0')

    # Ingreso proyectado (formula) — mix P1=30% P2=40% P3=20% Puerta=10%
    set_formula(ws_e, r, 8,
                f"=G{r}*(C{r}*0.30+D{r}*0.40+E{r}*0.20+F{r}*0.10)",
                fmt=CLP, bold=True)

    ent_rows.append(r)

# Total entradas
tot_e_row = 4 + len(tickets)
ws_e.row_dimensions[tot_e_row].height = 26
ws_e.merge_cells(f"A{tot_e_row}:G{tot_e_row}")
te = ws_e[f"A{tot_e_row}"]
te.value = "TOTAL ASISTENTES / INGRESOS PROYECTADOS POR ENTRADAS"
te.font = Font(name="Arial", bold=True, color=C_WHITE, size=11)
te.fill = fill(C_TEAL)
te.alignment = right_align()
te.border = med_border()

t_ing = ws_e.cell(row=tot_e_row, column=8)
t_ing.value = f"=SUM(H{ent_rows[0]}:H{ent_rows[-1]})"
t_ing.font = Font(name="Arial", bold=True, color=C_WHITE, size=12)
t_ing.fill = fill(C_TEAL)
t_ing.border = med_border()
t_ing.alignment = right_align()
t_ing.number_format = CLP

TOTAL_ENTRADAS_REF = f"'🎟️ Entradas'!H{tot_e_row}"

# Mix assumption note
note_row = tot_e_row + 2
ws_e.merge_cells(f"A{note_row}:H{note_row}")
ws_e[f"A{note_row}"].value = ("💡 Supuesto de mix de ventas aplicado al Ingreso Proyectado: "
                               "30% a precio Preventa 1 · 40% a Preventa 2 · 20% a Preventa 3 · 10% en Puerta. "
                               "Modifica los precios y asistentes según tu estimación real.")
ws_e[f"A{note_row}"].font = Font(name="Arial", color="333333", size=9, italic=True)
ws_e[f"A{note_row}"].alignment = left(wrap=True)
ws_e.row_dimensions[note_row].height = 30


# ════════════════════════════════════════════════════════════════════════════
#  HOJA 4: RESUMEN EJECUTIVO
# ════════════════════════════════════════════════════════════════════════════
ws_r = wb.create_sheet("📊 Resumen")
ws_r.sheet_view.showGridLines = False

for col, w in {"A": 36, "B": 22, "C": 22}.items():
    ws_r.column_dimensions[col].width = w

ws_r.row_dimensions[1].height = 45
ws_r.row_dimensions[2].height = 22

ws_r.merge_cells("A1:C1")
t = ws_r["A1"]
t.value = "📊 RESUMEN FINANCIERO — CONGRESO ELECTRÓNICO 2026"
t.font = Font(name="Arial", bold=True, color=C_WHITE, size=15)
t.fill = fill(C_NAVY)
t.alignment = center()

ws_r.merge_cells("A2:C2")
t2 = ws_r["A2"]
t2.value = "Este resumen se actualiza solo al modificar las hojas Gastos, Auspicios y Entradas"
t2.font = Font(name="Arial", color=C_WHITE, size=9)
t2.fill = fill(C_BLUE)
t2.alignment = center()

def add_kpi(ws, row, label, formula, color_bg, color_txt=C_WHITE, fmt=CLP, emoji=""):
    ws.row_dimensions[row].height = 32
    lc = ws.cell(row=row, column=1)
    lc.value = f"{emoji}  {label}"
    lc.font = Font(name="Arial", bold=True, color="222222", size=11)
    lc.fill = fill("F7F9FC")
    lc.border = thin_border()
    lc.alignment = left()

    vc = ws.cell(row=row, column=2)
    vc.value = formula
    vc.font = Font(name="Arial", bold=True, color=color_txt, size=13)
    vc.fill = fill(color_bg)
    vc.border = thin_border()
    vc.alignment = right_align()
    vc.number_format = fmt

    ws.merge_cells(f"B{row}:C{row}")

# ── Section: GASTOS ───────────────────────────────────────────────────────
r = 4
ws_r.row_dimensions[r].height = 26
ws_r.merge_cells(f"A{r}:C{r}")
sec = ws_r[f"A{r}"]
sec.value = "▶  GASTOS"
sec.font = Font(name="Arial", bold=True, color=C_WHITE, size=11)
sec.fill = fill(C_RED)
sec.alignment = center(); sec.border = thin_border()

r += 1
add_kpi(ws_r, r, "Total Presupuesto de Gastos",
        f"='💰 Gastos'!{TOTAL_GASTOS_REF}", C_RED, emoji="💸")

# ── Section: INGRESOS ─────────────────────────────────────────────────────
r += 2
ws_r.row_dimensions[r].height = 26
ws_r.merge_cells(f"A{r}:C{r}")
sec2 = ws_r[f"A{r}"]
sec2.value = "▶  INGRESOS"
sec2.font = Font(name="Arial", bold=True, color=C_WHITE, size=11)
sec2.fill = fill(C_TEAL)
sec2.alignment = center(); sec2.border = thin_border()

r += 1
add_kpi(ws_r, r, "Ingresos proyectados por entradas",
        f"={TOTAL_ENTRADAS_REF}", C_TEAL, emoji="🎟️")
r += 1
add_kpi(ws_r, r, "Auspicios acordados (total convenios)",
        f"={TOTAL_AUSPICIOS_ACORDADO_REF}", "D4AC0D", emoji="🤝")
r += 1
add_kpi(ws_r, r, "Auspicios recibidos (dinero en caja)",
        f"={TOTAL_AUSPICIOS_RECIBIDO_REF}", "1A7B6B", emoji="💵")

# ── Section: BALANCE ──────────────────────────────────────────────────────
r += 2
ws_r.row_dimensions[r].height = 26
ws_r.merge_cells(f"A{r}:C{r}")
sec3 = ws_r[f"A{r}"]
sec3.value = "▶  BALANCE Y NECESIDADES"
sec3.font = Font(name="Arial", bold=True, color=C_WHITE, size=11)
sec3.fill = fill(C_NAVY)
sec3.alignment = center(); sec3.border = thin_border()

r += 1
add_kpi(ws_r, r, "Total ingresos proyectados (entradas + auspicios acordados)",
        f"={TOTAL_ENTRADAS_REF}+{TOTAL_AUSPICIOS_ACORDADO_REF}",
        C_BLUE, emoji="📈")

r += 1
ws_r.row_dimensions[r].height = 32
lc = ws_r.cell(row=r, column=1)
lc.value = "💰  Balance neto (Ingresos proyectados − Gastos)"
lc.font = Font(name="Arial", bold=True, color="222222", size=11)
lc.fill = fill("F7F9FC"); lc.border = thin_border(); lc.alignment = left()
vc = ws_r.cell(row=r, column=2)
vc.value = (f"={TOTAL_ENTRADAS_REF}+{TOTAL_AUSPICIOS_ACORDADO_REF}"
            f"-'💰 Gastos'!{TOTAL_GASTOS_REF}")
vc.font = Font(name="Arial", bold=True, color=C_WHITE, size=13)
vc.fill = fill(C_NAVY); vc.border = thin_border()
vc.alignment = right_align(); vc.number_format = CLP
ws_r.merge_cells(f"B{r}:C{r}")
BALANCE_ROW = r

r += 1
add_kpi(ws_r, r, "⚠️  Monto aún necesario (solo con caja real recibida)",
        (f"='💰 Gastos'!{TOTAL_GASTOS_REF}"
         f"-{TOTAL_ENTRADAS_REF}"
         f"-{TOTAL_AUSPICIOS_RECIBIDO_REF}"),
        C_RED, emoji="🎯")

# Conditional note
r += 2
ws_r.row_dimensions[r].height = 20
ws_r.merge_cells(f"A{r}:C{r}")
ws_r[f"A{r}"].value = "▶  COBERTURA DEL PRESUPUESTO"
ws_r[f"A{r}"].font = Font(name="Arial", bold=True, color=C_WHITE, size=11)
ws_r[f"A{r}"].fill = fill(C_NAVY)
ws_r[f"A{r}"].alignment = center()
ws_r[f"A{r}"].border = thin_border()

r += 1
ws_r.row_dimensions[r].height = 28
lc2 = ws_r.cell(row=r, column=1)
lc2.value = "% del presupuesto cubierto (ingresos totales / gastos)"
lc2.font = Font(name="Arial", bold=True, color="222222", size=11)
lc2.fill = fill("F7F9FC"); lc2.border = thin_border(); lc2.alignment = left()
vc2 = ws_r.cell(row=r, column=2)
vc2.value = (f"=IFERROR(({TOTAL_ENTRADAS_REF}+{TOTAL_AUSPICIOS_ACORDADO_REF})"
             f"/'💰 Gastos'!{TOTAL_GASTOS_REF},0)")
vc2.font = Font(name="Arial", bold=True, color=C_WHITE, size=13)
vc2.fill = fill(C_TEAL); vc2.border = thin_border()
vc2.alignment = right_align(); vc2.number_format = "0.0%"
ws_r.merge_cells(f"B{r}:C{r}")

r += 1
ws_r.row_dimensions[r].height = 28
lc3 = ws_r.cell(row=r, column=1)
lc3.value = "% cubierto con dinero en caja real (auspicios recibidos + entradas)"
lc3.font = Font(name="Arial", bold=True, color="222222", size=11)
lc3.fill = fill("F7F9FC"); lc3.border = thin_border(); lc3.alignment = left()
vc3 = ws_r.cell(row=r, column=2)
vc3.value = (f"=IFERROR(({TOTAL_ENTRADAS_REF}+{TOTAL_AUSPICIOS_RECIBIDO_REF})"
             f"/'💰 Gastos'!{TOTAL_GASTOS_REF},0)")
vc3.font = Font(name="Arial", bold=True, color=C_WHITE, size=13)
vc3.fill = fill("D4A017"); vc3.border = thin_border()
vc3.alignment = right_align(); vc3.number_format = "0.0%"
ws_r.merge_cells(f"B{r}:C{r}")

r += 2
ws_r.row_dimensions[r].height = 36
ws_r.merge_cells(f"A{r}:C{r}")
hint = ws_r[f"A{r}"]
hint.value = ("💡 Instrucciones: Esta hoja no se edita directamente. "
              "Ve a 💰 Gastos para ajustar el presupuesto · "
              "Ve a 🤝 Auspicios para registrar auspiciadores · "
              "Ve a 🎟️ Entradas para ajustar precios y asistentes esperados.")
hint.font = Font(name="Arial", color="333333", size=9, italic=True)
hint.alignment = left(wrap=True)
ws_r.row_dimensions[r].height = 35

# ── Set tab order / colors ────────────────────────────────────────────────
ws_r.sheet_properties.tabColor = "2E4FA3"
ws_g.sheet_properties.tabColor = C_RED
ws_a.sheet_properties.tabColor = "D4A017"
ws_e.sheet_properties.tabColor = "1A7B6B"

# ── Move Resumen to first position ───────────────────────────────────────
wb.move_sheet("📊 Resumen", offset=-3)

# Save
out_path = "presupuesto_congreso_2026.xlsx"
wb.save(out_path)
print("Saved:", out_path)